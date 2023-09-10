import openpyxl
from pathlib import Path


def get_data(sheet_name):
    current_dir = Path().absolute().parent
    workbook = openpyxl.load_workbook(
        f"{current_dir}/Test_Data/appium_data.xlsx")  # This will work for MAC need to change for Windows
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    main_list = []

    for i in range(2, total_rows + 1):
        data_list = []
        for j in range(1, total_cols + 1):
            data = sheet.cell(row=i, column=j).value
            data_list.insert(j, data)
        main_list.insert(i, data_list)
    return main_list

